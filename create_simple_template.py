import openpyxl
from pathlib import Path
from datetime import datetime, timedelta

def create_simple_excel_template():
    """Pandas olmadan basit Excel şablonu oluştur"""
    
    print("🚀 Excel Şablonu Oluşturuluyor (Pandas-Free)")
    print("=" * 50)
    
    # Excel dosya yolu
    excel_file_path = Path("data/proje_sablonu.xlsx")
    excel_file_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Workbook oluştur
    wb = openpyxl.Workbook()
    
    # İlk sayfayı sil
    wb.remove(wb.active)
    
    # 1. GÖREVLER SAYFASI
    ws_tasks = wb.create_sheet("Görevler")
    
    # Başlıklar
    headers = [
        'Görev Adı', 'Süre (Gün)', 'Başlangıç Tarihi', 'Bağımlı Görevler', 
        'Atanan Kaynaklar', 'Görev Türü', 'Öncelik', 'Notlar'
    ]
    
    for col, header in enumerate(headers, 1):
        ws_tasks.cell(row=1, column=col, value=header)
    
    # Örnek görevler (basitleştirilmiş)
    tasks = [
        # Ana Kategoriler
        ["SALON ALANI", 42, "2025-07-28", "", "Proje Takımı", "Summary", "Yüksek", "En kritik alan"],
        ["  1.1 - Salon Montaj İşleri", 3, "2025-07-28", "1", "Kaynakçı[3]", "Normal", "Yüksek", "Saha hazırlığı"],
        ["  1.2 - Salon İmalat İşleri", 5, "2025-07-31", "2", "Kaynakçı[2]", "Normal", "Yüksek", "Kalite kontrolü"],
        
        ["FUAYE ALANI", 35, "2025-08-04", "", "Proje Takımı", "Summary", "Yüksek", "Salon ile paralel"],
        ["  2.1 - Fuaye Montaj İşleri", 3, "2025-08-04", "4", "Kaynakçı[2]", "Normal", "Orta", "Montaj sırası"],
        ["  2.2 - Fuaye İmalat İşleri", 4, "2025-08-07", "5", "Kaynakçı[1]", "Normal", "Orta", "Kalite kontrol"],
        
        ["SPOR SALONLARI", 45, "2025-08-11", "", "Proje Takımı", "Summary", "Yüksek", "Geniş alan"],
        ["  3.1 - Spor Montaj İşleri", 4, "2025-08-11", "7", "Kaynakçı[2]", "Normal", "Orta", "Paralel montaj"],
        ["  3.2 - Spor İmalat İşleri", 5, "2025-08-15", "8", "Kaynakçı[1]", "Normal", "Orta", "Hızlı imalat"],
        
        ["LOCALAR", 28, "2025-08-18", "", "Proje Takımı", "Summary", "Orta", "Küçük alan"],
        ["  4.1 - Loca Montaj İşleri", 2, "2025-08-18", "10", "Kaynakçı[1]", "Normal", "Orta", "Detay montaj"],
        ["  4.2 - Loca İmalat İşleri", 3, "2025-08-20", "11", "Kaynakçı[1]", "Normal", "Orta", "Hassas imalat"],
        
        ["TEKNİK OFİSLER", 30, "2025-08-25", "", "Proje Takımı", "Summary", "Orta", "Teknik alanlar"],
        ["  5.1 - Teknik Montaj İşleri", 2, "2025-08-25", "13", "Kaynakçı[1]", "Normal", "Orta", "Teknik montaj"],
        ["  5.2 - Teknik İmalat İşleri", 3, "2025-08-27", "14", "Kaynakçı[1]", "Normal", "Orta", "Hassas imalat"],
        
        ["ORTAK GÖREVLER", 25, "2025-09-22", "", "Proje Takımı", "Summary", "Yüksek", "Proje geneli"],
        ["  6.1 - Kediyolu İşleri", 3, "2025-09-22", "16", "Kaynakçı[2]", "Normal", "Yüksek", "Kediyolu imalatı"],
        ["  6.2 - Kalite Kontrol", 2, "2025-09-25", "17", "Teknisyen[2]", "Normal", "Yüksek", "Son kontroller"]
    ]
    
    # Görevleri ekle
    for row, task in enumerate(tasks, 2):
        for col, value in enumerate(task, 1):
            ws_tasks.cell(row=row, column=col, value=value)
    
    print(f"✅ {len(tasks)} görev eklendi")
    
    # 2. KAYNAKLAR SAYFASI
    ws_resources = wb.create_sheet("Kaynaklar")
    
    # Kaynak başlıkları
    resource_headers = ['Kaynak Adı', 'Tür', 'Maksimum Kullanım (%)', 'Birim Maliyet', 'Açıklama']
    for col, header in enumerate(resource_headers, 1):
        ws_resources.cell(row=1, column=col, value=header)
    
    # Kaynak verileri
    resources = [
        ['Proje Yöneticisi (Mimar)', 'İnsan', 100, 5000, 'Sorumlu mimar'],
        ['Usta Başı', 'İnsan', 100, 3500, 'Saha usta başı'],
        ['Kaynakçı-1', 'İnsan', 100, 2500, 'Birinci seviye kaynakçı'],
        ['Kaynakçı-2', 'İnsan', 100, 2500, 'Birinci seviye kaynakçı'],
        ['Kaynakçı-3', 'İnsan', 100, 2500, 'Birinci seviye kaynakçı'],
        ['Fitter-1', 'İnsan', 100, 2800, 'Çelik montaj uzmanı'],
        ['Fitter-2', 'İnsan', 100, 2800, 'Çelik montaj uzmanı'],
        ['26m Manlift-1', 'Ekipman', 100, 1500, '26 metre yükseklik kapasiteli'],
        ['26m Manlift-2', 'Ekipman', 100, 1500, '26 metre yükseklik kapasiteli'],
        ['Seyyar İskele-1', 'Ekipman', 100, 800, 'Taşınabilir çalışma platformu'],
        ['Kaynak Makinesi-1', 'Ekipman', 100, 300, 'MIG/MAG kaynak makinesi'],
        ['Kaynak Makinesi-2', 'Ekipman', 100, 300, 'MIG/MAG kaynak makinesi'],
        ['Teknisyen-1', 'İnsan', 100, 2200, 'Kontrol ve ölçüm uzmanı'],
        ['Teknisyen-2', 'İnsan', 100, 2200, 'Kontrol ve ölçüm uzmanı']
    ]
    
    # Kaynakları ekle
    for row, resource in enumerate(resources, 2):
        for col, value in enumerate(resource, 1):
            ws_resources.cell(row=row, column=col, value=value)
    
    print(f"✅ {len(resources)} kaynak eklendi")
    
    # 3. PROJE BİLGİLERİ SAYFASI
    ws_info = wb.create_sheet("Proje Bilgileri")
    
    # Proje bilgi başlıkları
    info_headers = ['Özellik', 'Değer', 'Açıklama']
    for col, header in enumerate(info_headers, 1):
        ws_info.cell(row=1, column=col, value=header)
    
    # Proje bilgileri
    project_info = [
        ['Proje Adı', 'Spor Salonu Çelik Takviye İşleri - 26.07.2025 Başlangıç', '26.07.2025 başlangıç ile 3 aylık optimize planlama'],
        ['Proje Yöneticisi', 'Taha Akgül', 'Sorumlu proje yöneticisi'],
        ['Başlangıç Tarihi', '2025-07-28', 'İlk iş gününden başlangıç (Pazartesi)'],
        ['Bitiş Tarihi', '2025-10-31', 'Son görevin tamamlanma tarihi'],
        ['Toplam Süre', '66 İş Günü (3 Ay)', 'Toplam çalışma süresi (hafta sonları hariç)'],
        ['Çalışma Günleri', 'Pazartesi-Cuma', 'Haftalık çalışma günleri'],
        ['Çalışma Saatleri', '08:00-17:00', 'Günlük çalışma saatleri'],
        ['Proje Durumu', 'Optimize Planlama - Paralel Çalışma', 'Paralel çalışma stratejisi ile optimize']
    ]
    
    # Proje bilgilerini ekle
    for row, info in enumerate(project_info, 2):
        for col, value in enumerate(info, 1):
            ws_info.cell(row=row, column=col, value=value)
    
    print(f"✅ Proje bilgileri eklendi")
    
    # Dosyayı kaydet
    wb.save(excel_file_path)
    
    print(f"✅ Excel şablonu oluşturuldu: {excel_file_path}")
    print(f"📅 Optimize Tarihi: 26.07.2025 başlangıç → 28.07.2025 (Pazartesi)")
    print(f"📅 Proje Bitiş Tarihi: 31.10.2025 (3 Ay)")
    print("\n📊 Şablon içeriği:")
    print(f"   • Görevler sayfası: {len(tasks)} görev")
    print(f"   • Kaynaklar sayfası: {len(resources)} kaynak")
    print("   • Proje Bilgileri sayfası: Optimize proje ayarları")
    print("\n🎯 Paralel Çalışma Stratejisi:")
    print("   • Salon Alanı: 28.07.2025 (Hemen başlar)")
    print("   • Fuaye Alanı: 04.08.2025 (1 hafta sonra)")
    print("   • Spor Salonları: 11.08.2025 (2 hafta sonra)")
    print("   • Localar: 18.08.2025 (3 hafta sonra)")
    print("   • Teknik Ofisler: 25.08.2025 (4 hafta sonra)")
    print("   • Ortak Görevler: 22.09.2025 (8 hafta sonra)")

if __name__ == "__main__":
    try:
        create_simple_excel_template()
        print("\n🎉 Başarıyla tamamlandı!")
    except Exception as e:
        print(f"❌ Hata: {e}")
        import traceback
        traceback.print_exc()
