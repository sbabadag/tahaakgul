#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MS Project Uyumlu Dosya Oluşturucu
Real MS Project format oluşturur - format hatası çözümü
"""

import os
import sys
from datetime import datetime, timedelta
import csv

def create_excel_to_msp_converter():
    """Excel'den MS Project'e gerçek dönüştürücü"""
    print("🔧 MS PROJECT UYUMLU DOSYA OLUŞTURUCU")
    print("=" * 50)
    
    try:
        # Excel dosyasını oku
        excel_file = "data/proje_sablonu.xlsx"
        if not os.path.exists(excel_file):
            print(f"❌ Excel dosyası bulunamadı: {excel_file}")
            return False
        
        print("📖 Excel dosyası okunuyor...")
        import openpyxl
        workbook = openpyxl.load_workbook(excel_file)
        
        if "Görevler" not in workbook.sheetnames:
            print("❌ Görevler sayfası bulunamadı!")
            return False
        
        tasks_sheet = workbook["Görevler"]
        
        # Görevleri oku
        tasks_data = []
        for row in tasks_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            
            task_info = {
                'id': row[0],
                'name': row[1] if row[1] else f"Görev {row[0]}",
                'duration': row[2] if row[2] else 1,
                'start': row[3] if row[3] else "28.07.2025",
                'finish': row[4] if len(row) > 4 and row[4] else None,
                'predecessors': row[5] if len(row) > 5 and row[5] else "",
                'resources': row[6] if len(row) > 6 and row[6] else "",
                'area': row[7] if len(row) > 7 and row[7] else "Genel",
                'priority': row[8] if len(row) > 8 and row[8] else "Orta"
            }
            tasks_data.append(task_info)
        
        print(f"✅ {len(tasks_data)} görev okundu")
        
        # Kaynakları oku
        resources_data = []
        if "Kaynaklar" in workbook.sheetnames:
            resources_sheet = workbook["Kaynaklar"]
            for row in resources_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                
                resource_info = {
                    'name': row[0],
                    'type': row[1] if row[1] else "Kaynak",
                    'cost': row[2] if row[2] else 0,
                    'max_units': row[3] if len(row) > 3 and row[3] else 100
                }
                resources_data.append(resource_info)
        
        print(f"✅ {len(resources_data)} kaynak okundu")
        
        # MS Project CSV formatı oluştur
        print("\n📄 MS Project CSV formatı oluşturuluyor...")
        csv_file = create_ms_project_csv(tasks_data, resources_data)
        
        # MS Project XML formatı oluştur
        print("📄 MS Project XML formatı oluşturuluyor...")
        xml_file = create_ms_project_xml(tasks_data, resources_data)
        
        # Kullanım talimatları
        print("\n" + "="*60)
        print("🎯 MS PROJECT'E AKTARIM TALİMATLARI")
        print("="*60)
        print("📂 Oluşturulan dosyalar:")
        print(f"   • CSV: {csv_file}")
        print(f"   • XML: {xml_file}")
        print()
        print("🔧 MS PROJECT'TE AÇMA:")
        print("1. Microsoft Project'i açın")
        print("2. Dosya > Aç > Türü: 'XML Files (*.xml)'")
        print(f"3. {xml_file} dosyasını seçin")
        print("4. Import Wizard açılacak:")
        print("   - 'New Map' seçin")
        print("   - 'Tasks' tabında eşleştirmeleri kontrol edin")
        print("   - 'Finish' tıklayın")
        print("5. Dosya > Farklı Kaydet > Tür: 'Project (*.mpp)'")
        print()
        print("🔄 ALTERNATİF: CSV İLE AÇMA:")
        print("1. Microsoft Project'i açın")
        print("2. Dosya > Aç > Türü: 'Text Files (*.txt, *.csv)'")
        print(f"3. {csv_file} dosyasını seçin")
        print("4. Import Wizard'da eşleştirmeleri yapın")
        print()
        
        return True
        
    except Exception as e:
        print(f"❌ Hata: {e}")
        import traceback
        traceback.print_exc()
        return False

def create_ms_project_csv(tasks_data, resources_data):
    """MS Project uyumlu CSV oluştur"""
    csv_file = "data/SporSalonu_MSProject_Compatible.csv"
    
    try:
        with open(csv_file, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            
            # CSV başlıkları - MS Project standart alanları
            headers = [
                'ID', 'Name', 'Duration', 'Start', 'Finish', 
                'Predecessors', 'Resource Names', 'Priority', 
                'Notes', 'Outline Level'
            ]
            writer.writerow(headers)
            
            # Görevleri yaz
            for task in tasks_data:
                # Süreyi MS Project formatına çevir
                duration = task['duration']
                if isinstance(duration, str):
                    duration = 1
                duration_str = f"{duration} days"
                
                # Tarihleri formatla
                start_date = task['start']
                if isinstance(start_date, str):
                    try:
                        dt = datetime.strptime(start_date, "%d.%m.%Y")
                        start_formatted = dt.strftime("%m/%d/%Y")
                    except:
                        start_formatted = "07/28/2025"
                else:
                    start_formatted = "07/28/2025"
                
                # Finish tarihi hesapla
                if task['finish']:
                    finish_date = task['finish']
                    if isinstance(finish_date, str):
                        try:
                            dt = datetime.strptime(finish_date, "%d.%m.%Y")
                            finish_formatted = dt.strftime("%m/%d/%Y")
                        except:
                            finish_formatted = ""
                    else:
                        finish_formatted = ""
                else:
                    finish_formatted = ""
                
                # Öncelik
                priority_map = {
                    "Düşük": "200",
                    "Orta": "500", 
                    "Yüksek": "800",
                    "Kritik": "1000"
                }
                priority = priority_map.get(task['priority'], "500")
                
                # Notlar
                notes = f"Alan: {task['area']}, Öncelik: {task['priority']}"
                
                # Satır yaz
                row = [
                    task['id'],
                    task['name'],
                    duration_str,
                    start_formatted,
                    finish_formatted,
                    task['predecessors'],
                    task['resources'],
                    priority,
                    notes,
                    "1"  # Outline level
                ]
                writer.writerow(row)
        
        print(f"   ✅ CSV oluşturuldu: {csv_file}")
        return csv_file
        
    except Exception as e:
        print(f"   ❌ CSV hatası: {e}")
        return None

def create_ms_project_xml(tasks_data, resources_data):
    """MS Project 2003/2007 uyumlu XML oluştur"""
    xml_file = "data/SporSalonu_MSProject_Compatible.xml"
    
    try:
        # MS Project XML namespace'i ile başla
        xml_content = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Project xmlns="http://schemas.microsoft.com/project">
    <Title>Spor Salonu Çelik Konstrüksiyon - Uyumlu Format</Title>
    <Company>Taha Akgül İnşaat</Company>
    <Manager>Proje Yöneticisi</Manager>
    <CreationDate>2025-07-21T08:00:00</CreationDate>
    <StartDate>2025-07-28T08:00:00</StartDate>
    <FinishDate>2025-10-31T17:00:00</FinishDate>
    <CurrencySymbol>TL</CurrencySymbol>
    <DefaultTaskType>1</DefaultTaskType>
    <DefaultFixedCostAccrual>3</DefaultFixedCostAccrual>
    <CalendarUID>1</CalendarUID>
    
    <!-- Takvim Tanımı -->
    <Calendars>
        <Calendar>
            <UID>1</UID>
            <Name>Standard</Name>
            <IsBaseCalendar>1</IsBaseCalendar>
            <WeekDays>
                <WeekDay>
                    <DayType>1</DayType>
                    <DayWorking>0</DayWorking>
                </WeekDay>
                <WeekDay>
                    <DayType>2</DayType>
                    <DayWorking>1</DayWorking>
                    <WorkingTimes>
                        <WorkingTime>
                            <FromTime>08:00:00</FromTime>
                            <ToTime>12:00:00</ToTime>
                        </WorkingTime>
                        <WorkingTime>
                            <FromTime>13:00:00</FromTime>
                            <ToTime>17:00:00</ToTime>
                        </WorkingTime>
                    </WorkingTimes>
                </WeekDay>
                <WeekDay>
                    <DayType>3</DayType>
                    <DayWorking>1</DayWorking>
                    <WorkingTimes>
                        <WorkingTime>
                            <FromTime>08:00:00</FromTime>
                            <ToTime>12:00:00</ToTime>
                        </WorkingTime>
                        <WorkingTime>
                            <FromTime>13:00:00</FromTime>
                            <ToTime>17:00:00</ToTime>
                        </WorkingTime>
                    </WorkingTimes>
                </WeekDay>
                <WeekDay>
                    <DayType>4</DayType>
                    <DayWorking>1</DayWorking>
                    <WorkingTimes>
                        <WorkingTime>
                            <FromTime>08:00:00</FromTime>
                            <ToTime>12:00:00</ToTime>
                        </WorkingTime>
                        <WorkingTime>
                            <FromTime>13:00:00</FromTime>
                            <ToTime>17:00:00</ToTime>
                        </WorkingTime>
                    </WorkingTimes>
                </WeekDay>
                <WeekDay>
                    <DayType>5</DayType>
                    <DayWorking>1</DayWorking>
                    <WorkingTimes>
                        <WorkingTime>
                            <FromTime>08:00:00</FromTime>
                            <ToTime>12:00:00</ToTime>
                        </WorkingTime>
                        <WorkingTime>
                            <FromTime>13:00:00</FromTime>
                            <ToTime>17:00:00</ToTime>
                        </WorkingTime>
                    </WorkingTimes>
                </WeekDay>
                <WeekDay>
                    <DayType>6</DayType>
                    <DayWorking>1</DayWorking>
                    <WorkingTimes>
                        <WorkingTime>
                            <FromTime>08:00:00</FromTime>
                            <ToTime>12:00:00</ToTime>
                        </WorkingTime>
                        <WorkingTime>
                            <FromTime>13:00:00</FromTime>
                            <ToTime>17:00:00</ToTime>
                        </WorkingTime>
                    </WorkingTimes>
                </WeekDay>
                <WeekDay>
                    <DayType>7</DayType>
                    <DayWorking>0</DayWorking>
                </WeekDay>
            </WeekDays>
        </Calendar>
    </Calendars>
    
    <!-- Kaynaklar -->
    <Resources>
'''
        
        # Kaynakları ekle
        for i, resource in enumerate(resources_data, 1):
            xml_content += f'''        <Resource>
            <UID>{i}</UID>
            <ID>{i}</ID>
            <Name>{resource['name']}</Name>
            <Type>1</Type>
            <IsNull>0</IsNull>
            <StandardRate>{resource['cost'] * 10000}</StandardRate>
            <MaxUnits>{resource['max_units'] / 100}</MaxUnits>
            <Group>{resource['type']}</Group>
        </Resource>
'''
        
        xml_content += '''    </Resources>
    
    <!-- Görevler -->
    <Tasks>
'''
        
        # Görevleri ekle
        for i, task in enumerate(tasks_data, 1):
            # Süre hesapla (dakika cinsinden)
            duration = task['duration'] if task['duration'] else 1
            if isinstance(duration, str):
                try:
                    duration = int(duration.replace('d', '').strip())
                except:
                    duration = 1
            
            duration_minutes = duration * 8 * 60  # gün * 8 saat * 60 dakika
            
            # Başlangıç tarihini formatla
            start_date = task['start']
            if isinstance(start_date, str):
                try:
                    dt = datetime.strptime(start_date, "%d.%m.%Y")
                    start_formatted = dt.strftime("%Y-%m-%dT08:00:00")
                except:
                    start_formatted = "2025-07-28T08:00:00"
            else:
                start_formatted = "2025-07-28T08:00:00"
            
            # Öncelik
            priority_map = {
                "Düşük": "200",
                "Orta": "500",
                "Yüksek": "800", 
                "Kritik": "1000"
            }
            priority = priority_map.get(task['priority'], "500")
            
            xml_content += f'''        <Task>
            <UID>{i}</UID>
            <ID>{i}</ID>
            <Name>{task['name']}</Name>
            <Type>1</Type>
            <IsNull>0</IsNull>
            <CreateDate>2025-07-21T08:00:00</CreateDate>
            <WBS>{i}</WBS>
            <OutlineLevel>1</OutlineLevel>
            <Priority>{priority}</Priority>
            <Start>{start_formatted}</Start>
            <Duration>PT{duration_minutes}M</Duration>
            <DurationFormat>7</DurationFormat>
            <Work>PT{duration_minutes}M</Work>
            <Notes>Alan: {task['area']}, Öncelik: {task['priority']}</Notes>
            <ConstraintType>0</ConstraintType>
            <CalendarUID>1</CalendarUID>
'''
            
            # Bağımlılıklar varsa ekle
            if task['predecessors']:
                predecessors = str(task['predecessors']).split(',')
                for pred in predecessors:
                    pred = pred.strip()
                    if pred:
                        xml_content += f'''            <PredecessorLink>
                <PredecessorUID>{pred}</PredecessorUID>
                <Type>1</Type>
            </PredecessorLink>
'''
            
            xml_content += '''        </Task>
'''
        
        xml_content += '''    </Tasks>
</Project>'''
        
        # XML dosyasını yaz
        with open(xml_file, 'w', encoding='utf-8') as file:
            file.write(xml_content)
        
        print(f"   ✅ XML oluşturuldu: {xml_file}")
        return xml_file
        
    except Exception as e:
        print(f"   ❌ XML hatası: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """Ana işlem"""
    print("🚀 MS PROJECT UYUMLU DOSYA OLUŞTURUCU")
    print("Format Hatası Çözümü - Real MS Project Files")
    print("=" * 60)
    
    success = create_excel_to_msp_converter()
    
    if success:
        print("\n🎉 MS PROJECT UYUMLU DOSYALAR OLUŞTURULDU!")
        print("💡 Yukarıdaki talimatları takip ederek MS Project'te açın")
        
        # Mevcut MPP dosyasını sil (format hatası veren)
        old_mpp = "data/SporSalonu_Optimized_26_07_2025.mpp"
        if os.path.exists(old_mpp):
            try:
                os.remove(old_mpp)
                print(f"🗑️ Eski format hatası veren dosya silindi: {old_mpp}")
            except:
                print(f"⚠️ Eski dosya silinemedi: {old_mpp}")
    else:
        print("\n❌ Dosya oluşturulamadı!")
    
    return success

if __name__ == "__main__":
    success = main()
    if not success:
        sys.exit(1)
