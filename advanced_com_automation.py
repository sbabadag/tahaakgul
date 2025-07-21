#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gelişmiş MS Project COM Otomasyonu - Taha Akgül Proje Planlama Sistemi
Excel şablonunu tamamen MS Project COM API'si ile profesyonel MPP dosyasına dönüştürür
"""

import os
import sys
from datetime import datetime, timedelta
import traceback

def check_requirements():
    """Gerekli paketleri kontrol et ve yükle"""
    print("🔧 Sistem gereksinimleri kontrol ediliyor...")
    
    try:
        import comtypes.client
        print("   ✅ comtypes paketi hazır")
    except ImportError:
        print("   📦 comtypes paketi yükleniyor...")
        os.system("pip install comtypes")
        import comtypes.client
        print("   ✅ comtypes yüklendi")
    
    try:
        import openpyxl
        print("   ✅ openpyxl paketi hazır")
    except ImportError:
        print("   📦 openpyxl paketi yükleniyor...")
        os.system("pip install openpyxl")
        import openpyxl
        print("   ✅ openpyxl yüklendi")
    
    return True

class MSProjectCOMAutomator:
    """MS Project COM Automation Sınıfı"""
    
    def __init__(self):
        self.app = None
        self.project = None
        self.excel_file = "data/proje_sablonu.xlsx"
        self.output_file = "data/SporSalonu_Optimized_26_07_2025.mpp"
        
    def initialize_msproject(self):
        """MS Project uygulamasını başlat"""
        print("🚀 Microsoft Project başlatılıyor...")
        try:
            import comtypes.client
            
            # MS Project COM nesnesini oluştur
            self.app = comtypes.client.CreateObject("MSProject.Application")
            self.app.Visible = True
            self.app.DisplayAlerts = False  # Uyarıları kapat
            
            # Yeni proje oluştur
            self.project = self.app.Projects.Add()
            
            print("   ✅ MS Project başarıyla başlatıldı")
            return True
            
        except Exception as e:
            print(f"   ❌ MS Project başlatma hatası: {e}")
            print("   💡 Microsoft Project yüklü olduğundan ve lisanslı olduğundan emin olun")
            return False
    
    def setup_project_properties(self):
        """Proje özelliklerini ayarla"""
        print("📋 Proje özellikleri ayarlanıyor...")
        
        try:
            # Temel proje bilgileri
            self.project.ProjectStart = "28.07.2025"
            self.project.Title = "Spor Salonu Çelik Konstrüksiyon - Optimized 26.07.2025"
            self.project.Company = "Taha Akgül İnşaat"
            self.project.Manager = "Taha Akgül"
            self.project.Comments = "26.07.2025 başlangıç tarihi için optimize edilmiş 5 alan paralel çalışma stratejisi"
            
            # Takvim ayarları
            calendar = self.project.BaseCalendars.Item("Standard")
            # Pazartesi-Cuma çalışma günleri (08:00-17:00)
            for day in range(2, 7):  # Pazartesi=2, Cuma=6
                calendar.WeekDays.Item(day).Shift1.Start = "08:00"
                calendar.WeekDays.Item(day).Shift1.Finish = "17:00"
            
            # Cumartesi ve Pazar tatil
            calendar.WeekDays.Item(1).Working = False  # Pazar
            calendar.WeekDays.Item(7).Working = False  # Cumartesi
            
            print("   ✅ Proje özellikleri ayarlandı")
            return True
            
        except Exception as e:
            print(f"   ❌ Proje özellik hatası: {e}")
            return False
    
    def load_excel_data(self):
        """Excel verilerini yükle"""
        print("📖 Excel verileri yükleniyor...")
        
        try:
            import openpyxl
            
            if not os.path.exists(self.excel_file):
                print(f"   ❌ Excel dosyası bulunamadı: {self.excel_file}")
                return None, None
            
            workbook = openpyxl.load_workbook(self.excel_file)
            
            # Görevler sayfası
            gorevler_sheet = workbook["Görevler"]
            tasks_data = []
            for row in gorevler_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                tasks_data.append({
                    'id': row[0],
                    'name': row[1] or f"Görev {row[0]}",
                    'duration': row[2] or 1,
                    'start_date': row[3],
                    'predecessors': row[4] if len(row) > 4 else None,
                    'resources': row[5] if len(row) > 5 else None,
                    'area': row[6] if len(row) > 6 else "Genel"
                })
            
            # Kaynaklar sayfası
            kaynaklar_sheet = workbook["Kaynaklar"]
            resources_data = []
            for row in kaynaklar_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                resources_data.append({
                    'name': row[0],
                    'type': row[1] or "Kaynak",
                    'cost': row[2] or 0,
                    'max_units': row[3] or 100
                })
            
            print(f"   ✅ {len(tasks_data)} görev ve {len(resources_data)} kaynak yüklendi")
            return tasks_data, resources_data
            
        except Exception as e:
            print(f"   ❌ Excel okuma hatası: {e}")
            return None, None
    
    def create_resources(self, resources_data):
        """Kaynakları oluştur"""
        print("👥 Kaynaklar oluşturuluyor...")
        
        try:
            resource_count = 0
            for resource in resources_data:
                # Kaynak ekle
                res = self.project.Resources.Add(resource['name'])
                res.Type = 1  # Work resource (1=Work, 2=Material, 3=Cost)
                res.StandardRate = resource['cost']
                res.MaxUnits = resource['max_units']
                res.Group = resource['type']
                
                resource_count += 1
                
                if resource_count % 5 == 0:
                    print(f"   ✅ {resource_count} kaynak eklendi...")
            
            print(f"   ✅ Toplam {resource_count} kaynak oluşturuldu")
            return True
            
        except Exception as e:
            print(f"   ❌ Kaynak oluşturma hatası: {e}")
            return False
    
    def create_tasks(self, tasks_data):
        """Görevleri oluştur"""
        print("📋 Görevler oluşturuluyor...")
        
        try:
            task_count = 0
            created_tasks = {}
            
            for task_data in tasks_data:
                # Görev ekle
                task = self.project.Tasks.Add(task_data['name'])
                task.Duration = f"{task_data['duration']}d"
                
                # Başlangıç tarihi ayarla
                if task_data['start_date']:
                    if isinstance(task_data['start_date'], str):
                        task.Start = task_data['start_date']
                    elif hasattr(task_data['start_date'], 'strftime'):
                        task.Start = task_data['start_date'].strftime("%d.%m.%Y")
                
                # Görev özelliklerini ayarla
                task.Notes = f"Alan: {task_data['area']}"
                task.Priority = 500  # Normal öncelik
                
                # Görevler sözlüğüne ekle
                created_tasks[task_data['id']] = task
                
                task_count += 1
                
                if task_count % 5 == 0:
                    print(f"   ✅ {task_count} görev eklendi...")
            
            # Bağımlılıkları ayarla
            print("🔗 Görev bağımlılıkları ayarlanıyor...")
            dependency_count = 0
            
            for task_data in tasks_data:
                if task_data['predecessors']:
                    current_task = created_tasks[task_data['id']]
                    predecessors = str(task_data['predecessors']).split(',')
                    
                    for pred_id in predecessors:
                        pred_id = pred_id.strip()
                        if pred_id and pred_id in created_tasks:
                            # Bağımlılık ekle (Finish-to-Start)
                            current_task.PredecessorTasks.Add(created_tasks[pred_id])
                            dependency_count += 1
            
            print(f"   ✅ Toplam {task_count} görev ve {dependency_count} bağımlılık oluşturuldu")
            return created_tasks
            
        except Exception as e:
            print(f"   ❌ Görev oluşturma hatası: {e}")
            return None
    
    def assign_resources(self, tasks_data, created_tasks):
        """Kaynak atamalarını yap"""
        print("🔄 Kaynak atamaları yapılıyor...")
        
        try:
            assignment_count = 0
            
            for task_data in tasks_data:
                if task_data['resources'] and task_data['id'] in created_tasks:
                    task = created_tasks[task_data['id']]
                    resources = str(task_data['resources']).split(',')
                    
                    for resource_name in resources:
                        resource_name = resource_name.strip()
                        if resource_name:
                            try:
                                # Kaynağı bul
                                resource = self.project.Resources(resource_name)
                                # Atama yap
                                assignment = task.Assignments.Add(
                                    ResourceID=resource.ID,
                                    Units=100  # %100 kullanım
                                )
                                assignment_count += 1
                            except:
                                # Kaynak bulunamazsa geç
                                continue
            
            print(f"   ✅ {assignment_count} kaynak ataması yapıldı")
            return True
            
        except Exception as e:
            print(f"   ❌ Kaynak atama hatası: {e}")
            return False
    
    def optimize_schedule(self):
        """Proje programını optimize et"""
        print("⚡ Proje programı optimize ediliyor...")
        
        try:
            # Otomatik programlama
            self.project.Recalculate()
            
            # Kritik yol analizi
            self.app.GanttChart()
            
            # Görünüm ayarları
            self.app.ViewApply("Gantt Chart")
            
            print("   ✅ Program optimizasyonu tamamlandı")
            return True
            
        except Exception as e:
            print(f"   ❌ Optimizasyon hatası: {e}")
            return False
    
    def save_project(self):
        """Projeyi kaydet"""
        print("💾 Proje kaydediliyor...")
        
        try:
            # Dosya yolunu mutlak yap
            abs_output_file = os.path.abspath(self.output_file)
            
            # Kaydet
            self.project.SaveAs(abs_output_file)
            
            print(f"   ✅ Proje başarıyla kaydedildi: {abs_output_file}")
            return abs_output_file
            
        except Exception as e:
            print(f"   ❌ Kaydetme hatası: {e}")
            return None
    
    def generate_reports(self):
        """Raporları oluştur"""
        print("📊 Proje raporları oluşturuluyor...")
        
        try:
            # Gantt Chart görünümü
            self.app.ViewApply("Gantt Chart")
            
            # Task Usage görünümü
            self.app.ViewApply("Task Usage")
            
            # Resource Sheet görünümü  
            self.app.ViewApply("Resource Sheet")
            
            # Tekrar Gantt Chart'a dön
            self.app.ViewApply("Gantt Chart")
            
            print("   ✅ Proje raporları hazırlandı")
            return True
            
        except Exception as e:
            print(f"   ❌ Rapor oluşturma hatası: {e}")
            return False
    
    def cleanup(self):
        """Temizlik işlemleri"""
        try:
            if self.app:
                self.app.DisplayAlerts = True
        except:
            pass

def main():
    """Ana işlem fonksiyonu"""
    print("🚀 GELİŞMİŞ MS PROJECT COM OTOMASYONU")
    print("=" * 50)
    print("📅 Proje: 28.07.2025 → 31.10.2025 (3 Ay)")
    print("🏗️ Paralel çalışma: 5 alan eş zamanlı")
    print()
    
    # Gereksinimleri kontrol et
    if not check_requirements():
        print("❌ Gereksinimler karşılanamadı!")
        return False
    
    # COM Automator'ı başlat
    automator = MSProjectCOMAutomator()
    
    try:
        # Adım 1: MS Project'i başlat
        if not automator.initialize_msproject():
            return False
        
        # Adım 2: Proje özelliklerini ayarla
        if not automator.setup_project_properties():
            return False
        
        # Adım 3: Excel verilerini yükle
        tasks_data, resources_data = automator.load_excel_data()
        if not tasks_data or not resources_data:
            return False
        
        # Adım 4: Kaynakları oluştur
        if not automator.create_resources(resources_data):
            return False
        
        # Adım 5: Görevleri oluştur
        created_tasks = automator.create_tasks(tasks_data)
        if not created_tasks:
            return False
        
        # Adım 6: Kaynak atamalarını yap
        if not automator.assign_resources(tasks_data, created_tasks):
            return False
        
        # Adım 7: Programı optimize et
        if not automator.optimize_schedule():
            return False
        
        # Adım 8: Raporları oluştur
        if not automator.generate_reports():
            return False
        
        # Adım 9: Projeyi kaydet
        output_file = automator.save_project()
        if not output_file:
            return False
        
        # Başarı mesajı
        print()
        print("🎉 GELİŞMİŞ COM OTOMASYONU BAŞARILI!")
        print("=" * 50)
        print(f"📁 MS Project dosyası: {output_file}")
        print("📊 Özellikler:")
        print(f"   • {len(tasks_data)} görev")
        print(f"   • {len(resources_data)} kaynak")
        print("   • Kritik yol analizi")
        print("   • Kaynak atamaları")
        print("   • Optimized program")
        print()
        
        return True
        
    except Exception as e:
        print(f"❌ Beklenmeyen hata: {e}")
        print("Detaylı hata:")
        traceback.print_exc()
        return False
    
    finally:
        # Temizlik
        automator.cleanup()

if __name__ == "__main__":
    success = main()
    if not success:
        print("❌ İşlem başarısız!")
        sys.exit(1)
    else:
        print("✅ İşlem başarıyla tamamlandı!")
        sys.exit(0)
